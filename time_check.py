import streamlit as st
import pandas as pd
from datetime import timedelta, datetime, time
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import plotly.express as px
import plotly.graph_objects as go
from io import BytesIO

st.set_page_config(page_title="时间间隔缺失检测与填充", page_icon="⏱️", layout="wide")
st.title("⏱️ 时间间隔缺失检测与填充工具")
st.markdown("上传Excel文件，自动检测异常时间间隔，线性插值填充缺失值，查看每日统计，下载标黄结果")

# ==================== 辅助函数 ====================
def safe_convert_for_display(df):
    """将DataFrame中的日期时间列转换为字符串"""
    df = df.copy()
    for col in df.columns:
        try:
            sample = df[col].dropna()
            if len(sample) == 0:
                continue
            val = sample.iloc[0]
            if isinstance(val, (pd.Timestamp, datetime)):
                df[col] = df[col].astype(str)
            elif isinstance(val, time):
                df[col] = df[col].apply(lambda x: x.strftime('%H:%M:%S') if isinstance(x, time) else str(x))
        except:
            pass
    return df

# ==================== 核心函数 ====================
def detect_time_columns(df):
    date_col, time_col, combined_col = None, None, None
    for col in df.columns:
        sample = df[col].dropna().iloc[0] if len(df[col].dropna()) > 0 else None
        if sample is None:
            continue
        if isinstance(sample, (datetime, pd.Timestamp)):
            has_time = any(t.hour != 0 or t.minute != 0 or t.second != 0 for t in df[col].dropna() if hasattr(t, 'hour'))
            has_date = any(t.year > 2000 for t in df[col].dropna() if hasattr(t, 'year'))
            if has_date and has_time:
                combined_col = col
            elif has_date:
                date_col = col
        if isinstance(sample, time):
            time_col = col
        if isinstance(sample, str):
            try:
                converted = pd.to_datetime(df[col], errors='coerce')
                if converted.notna().sum() / len(df) > 0.8:
                    has_time = any(t.hour != 0 or t.minute != 0 or t.second != 0 for t in converted.dropna())
                    has_date = any(t.year > 2000 for t in converted.dropna())
                    if has_date and has_time:
                        combined_col = col
                    elif has_date:
                        date_col = col
            except:
                try:
                    pd.to_datetime('2024-01-01 ' + sample)
                    time_col = col
                except:
                    pass
    return date_col, time_col, combined_col

def build_datetime_series(df, date_col, time_col, combined_col):
    if combined_col:
        return pd.to_datetime(df[combined_col], errors='coerce')
    elif date_col and time_col:
        if isinstance(df[time_col].iloc[0], time):
            time_str = df[time_col].apply(lambda t: t.strftime('%H:%M:%S'))
        else:
            time_str = df[time_col].astype(str).str.replace('：', ':')
        date_str = pd.to_datetime(df[date_col]).dt.strftime('%Y-%m-%d')
        return pd.to_datetime(date_str + ' ' + time_str, errors='coerce')
    return None

def analyze_and_fill(df, date_col, time_col, combined_col, value_cols, freq_minutes=10):
    datetime_series = build_datetime_series(df, date_col, time_col, combined_col)
    if datetime_series is None:
        st.error("无法识别时间列")
        return None, None, None
    df['_temp_time'] = datetime_series
    df = df.dropna(subset=['_temp_time']).reset_index(drop=True)
    df['_row_id'] = range(len(df))
    df['_is_original'] = True
    gaps = []
    for i in range(len(df) - 1):
        t1 = df.loc[i, '_temp_time']
        t2 = df.loc[i + 1, '_temp_time']
        delta = (t2 - t1).total_seconds() / 60.0
        if delta < 0:
            delta = abs(delta)
        missing = max(0, round(delta / freq_minutes) - 1)
        gaps.append({'index': i, 'time1': t1, 'time2': t2, 'delta': delta, 'missing_count': missing})
    gaps_df = pd.DataFrame(gaps)
    need_fill = gaps_df[gaps_df['missing_count'] > 0]
    if len(need_fill) == 0:
        df = df.drop(columns=['_temp_time', '_row_id', '_is_original'])
        return df, gaps_df, datetime_series
    new_rows = []
    for _, gap in need_fill.iterrows():
        i = int(gap['index'])
        missing_count = int(gap['missing_count'])
        t1 = df.loc[i, '_temp_time']
        t2 = df.loc[i + 1, '_temp_time']
        base_row_id = df.loc[i, '_row_id']
        for j in range(1, missing_count + 1):
            new_time = t1 + timedelta(minutes=freq_minutes * j)
            new_row = {'_row_id': base_row_id + j / (missing_count + 1), '_temp_time': new_time, '_is_original': False}
            if date_col and time_col:
                new_row[date_col] = new_time.date()
                new_row[time_col] = new_time.time()
            elif combined_col:
                new_row[combined_col] = new_time
            for col in value_cols:
                if col in df.columns:
                    v1 = df.loc[i, col]
                    v2 = df.loc[i + 1, col]
                    if pd.notna(v1) and pd.notna(v2):
                        frac = j / (missing_count + 1)
                        interpolated = v1 + (v2 - v1) * frac
                        if abs(interpolated) < 10:
                            new_row[col] = round(interpolated, 2)
                        elif abs(interpolated) < 100:
                            new_row[col] = round(interpolated, 1)
                        else:
                            new_row[col] = round(interpolated, 0)
                    else:
                        new_row[col] = np.nan
            for col in df.columns:
                if col not in new_row and col not in ['_row_id', '_temp_time', '_is_original']:
                    new_row[col] = np.nan
            new_rows.append(new_row)
    df_filled = pd.concat([df, pd.DataFrame(new_rows)], ignore_index=True)
    df_filled = df_filled.sort_values('_row_id').reset_index(drop=True)
    df_filled = df_filled.drop(columns=['_temp_time', '_row_id'])
    original_cols = [col for col in df.columns if col not in ['_temp_time', '_row_id', '_is_original']]
    df_filled = df_filled[original_cols + ['_is_original']]
    return df_filled, gaps_df, datetime_series

def calculate_daily_stats(df, date_col, combined_col, value_cols):
    if combined_col:
        df['_date'] = pd.to_datetime(df[combined_col]).dt.date
    elif date_col:
        df['_date'] = pd.to_datetime(df[date_col]).dt.date
    else:
        return None
    numeric_cols = [c for c in value_cols if c in df.columns and df[c].dtype in ['float64', 'int64', 'float32', 'int32']]
    if not numeric_cols:
        return None
    daily_stats = df.groupby('_date')[numeric_cols].mean().reset_index()
    daily_stats['_date'] = pd.to_datetime(daily_stats['_date'])
    return daily_stats

def to_excel_with_highlight(df, highlight_col='_is_original', highlight_color='FFFF00'):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active
    fill = PatternFill(start_color=highlight_color, end_color=highlight_color, fill_type="solid")
    highlight_col_idx = None
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value == highlight_col:
            highlight_col_idx = col_idx
            break
    if highlight_col_idx:
        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=highlight_col_idx).value == False:
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ==================== 颜色预设 ====================
COLOR_PRESETS = {
    "🟡 黄色": "FFFF00", "🟠 橙色": "FFA500", "🟢 浅绿": "90EE90",
    "🔴 浅红": "FFB6C1", "🔵 浅蓝": "ADD8E6", "🟣 浅紫": "DDA0DD",
    "⚪ 浅灰": "D3D3D3", "🟤 浅棕": "DEB887",
}

# ==================== 侧边栏 ====================
with st.sidebar:
    st.header("⚙️ 配置")
    uploaded_file = st.file_uploader("📤 上传Excel文件", type=['xlsx', 'xls'])
    
    if uploaded_file:
        df_original = pd.read_excel(uploaded_file)
        st.markdown("---")
        st.subheader("🕐 时间设置")
        freq_minutes = st.number_input("正常间隔（分钟）", min_value=1, value=10)
        date_col, time_col, combined_col = detect_time_columns(df_original)
        all_cols = df_original.columns.tolist()
        
        if combined_col:
            combined_col = st.selectbox("合并时间列", all_cols, index=all_cols.index(combined_col) if combined_col in all_cols else 0)
        
        date_col_options = ["（自动选择）"] + all_cols
        time_col_options = ["（自动选择）"] + all_cols
        
        date_idx = all_cols.index(date_col) + 1 if date_col and date_col in all_cols else 0
        time_idx = all_cols.index(time_col) + 1 if time_col and time_col in all_cols else 0
        
        date_col_sel = st.selectbox("日期列", date_col_options, index=date_idx)
        time_col_sel = st.selectbox("时间列", time_col_options, index=time_idx)
        
        date_col = None if date_col_sel == "（自动选择）" else date_col_sel
        time_col = None if time_col_sel == "（自动选择）" else time_col_sel
        
        st.markdown("---")
        st.subheader("📊 特征列设置")
        
        time_related = [c for c in [date_col, time_col, combined_col] if c]
        numeric_cols = [c for c in all_cols if c not in time_related and df_original[c].dtype in ['float64', 'int64', 'float32', 'int32']]
        default_cols = [c for c in numeric_cols][:min(4, len(numeric_cols))]
        
        value_cols = st.multiselect("选择特征列", options=all_cols, default=default_cols)
        
        manual_cols = st.text_input("手动输入列名（逗号分隔）", placeholder="例如: 温度, 湿度")
        if manual_cols:
            for c in manual_cols.split(','):
                c = c.strip()
                if c in all_cols and c not in value_cols:
                    value_cols.append(c)
                elif c not in all_cols:
                    st.warning(f"列 '{c}' 不存在")
        
        st.markdown("---")
        st.subheader("🎨 高亮颜色")
        color_name = st.selectbox("选择插入行颜色", list(COLOR_PRESETS.keys()), index=0)
        highlight_color = COLOR_PRESETS[color_name]
        st.markdown(f'<div style="background-color:#{highlight_color};padding:10px;border-radius:5px;text-align:center;">预览颜色</div>', unsafe_allow_html=True)

# ==================== 主界面 ====================
if uploaded_file is None:
    st.info("👈 请先在左侧上传Excel文件")
else:
    tab1, tab2, tab3, tab4, tab5 = st.tabs(["📋 原始数据", "📊 间隔分析", "🔧 填充结果", "📈 每日统计", "💾 下载"])
    
    with tab1:
        st.subheader("原始数据预览")
        st.dataframe(safe_convert_for_display(df_original.head(100)), use_container_width=True)
        c1, c2, c3 = st.columns(3)
        c1.metric("总行数", len(df_original))
        c2.metric("总列数", len(df_original.columns))
        c3.metric("特征列数", len(value_cols) if value_cols else 0)
    
    if value_cols:
        with st.spinner("正在分析..."):
            df_filled, gaps_df, datetime_series = analyze_and_fill(df_original.copy(), date_col, time_col, combined_col, value_cols, freq_minutes)
        
        if df_filled is not None and gaps_df is not None:
            normal_mask = (gaps_df['delta'] >= freq_minutes - 1) & (gaps_df['delta'] <= freq_minutes + 1)
            need_fill = gaps_df[gaps_df['missing_count'] > 0]
            total_missing = int(need_fill['missing_count'].sum()) if len(need_fill) > 0 else 0
            
            with tab2:
                st.subheader("📊 时间间隔分析")
                c1, c2, c3, c4, c5 = st.columns(5)
                c1.metric("总间隔数", len(gaps_df))
                c2.metric("✅ 正常", normal_mask.sum())
                c3.metric("⚠️ 异常", len(need_fill))
                c4.metric("📌 需插入", total_missing)
                pct = normal_mask.sum() / len(gaps_df) * 100 if len(gaps_df) > 0 else 0
                c5.metric("正常率", f"{pct:.1f}%")
                
                c_left, c_right = st.columns(2)
                with c_left:
                    fig = px.histogram(gaps_df, x='delta', nbins=50, labels={'delta': '间隔（分钟）', 'count': '数量'})
                    fig.add_vline(x=freq_minutes, line_dash="dash", line_color="green")
                    st.plotly_chart(fig, use_container_width=True)
                with c_right:
                    gaps_df['time_point'] = gaps_df['time1']
                    fig2 = px.scatter(gaps_df, x='time_point', y='delta', labels={'time_point': '时间', 'delta': '间隔（分钟）'})
                    fig2.add_hline(y=freq_minutes, line_dash="dash", line_color="green")
                    st.plotly_chart(fig2, use_container_width=True)
                
                if len(need_fill) > 0:
                    st.subheader(f"⚠️ 异常间隔详情（共{len(need_fill)}个）")
                    detail_df = need_fill[['index', 'time1', 'time2', 'delta', 'missing_count']].copy()
                    detail_df.columns = ['起始行', '开始时间', '结束时间', '间隔(分钟)', '需插入点数']
                    detail_df['间隔(分钟)'] = detail_df['间隔(分钟)'].round(1)
                    detail_df['需插入点数'] = detail_df['需插入点数'].astype(int)
                    st.dataframe(detail_df, use_container_width=True)
                else:
                    st.success("✅ 所有时间间隔正常")
            
            with tab3:
                st.subheader("🔧 填充结果")
                inserted_count = len(df_filled) - len(df_original)
                st.markdown(f'<div style="display:flex;align-items:center;margin-bottom:10px;"><div style="background-color:#{highlight_color};width:20px;height:20px;margin-right:10px;border:1px solid #ccc;"></div><span>= 插入行（共{inserted_count}行）</span></div>', unsafe_allow_html=True)
                
                if inserted_count > 0:
                    inserted_indices = df_filled[~df_filled['_is_original']].index.tolist()
                    display_indices = set()
                    for idx in inserted_indices:
                        for offset in range(-3, 4):
                            if 0 <= idx + offset < len(df_filled):
                                display_indices.add(idx + offset)
                    display_df = df_filled.iloc[sorted(display_indices)].copy()
                    display_df = safe_convert_for_display(display_df)
                    
                    def highlight_inserted(row):
                        if row.get('_is_original', True) == False:
                            return [f'background-color: #{highlight_color}'] * len(row)
                        return [''] * len(row)
                    
                    st.dataframe(display_df.style.apply(highlight_inserted, axis=1), use_container_width=True, height=500)
                    
                    with st.expander("查看所有插入行"):
                        inserted_df = df_filled[~df_filled['_is_original']].drop(columns=['_is_original'])
                        st.dataframe(safe_convert_for_display(inserted_df), use_container_width=True)
                else:
                    st.dataframe(safe_convert_for_display(df_filled.head(50)), use_container_width=True)
                
                c1, c2, c3 = st.columns(3)
                c1.metric("原始行数", len(df_original))
                c2.metric("填充后行数", len(df_filled))
                c3.metric("新增行数", inserted_count)
            
            with tab4:
                st.subheader("📈 每日特征均值")
                daily_stats = calculate_daily_stats(df_filled, date_col, combined_col, value_cols)
                if daily_stats is not None:
                    min_date = daily_stats['_date'].min()
                    max_date = daily_stats['_date'].max()
                    date_range = st.date_input("选择日期范围", value=(min_date, max_date), min_value=min_date, max_value=max_date)
                    if len(date_range) == 2:
                        mask = (daily_stats['_date'] >= pd.Timestamp(date_range[0])) & (daily_stats['_date'] <= pd.Timestamp(date_range[1]))
                        filtered_stats = daily_stats[mask]
                    else:
                        filtered_stats = daily_stats
                    st.dataframe(filtered_stats.style.format({c: '{:.2f}' for c in value_cols if c in filtered_stats.columns}), use_container_width=True)
                    
                    plot_cols = [c for c in value_cols if c in filtered_stats.columns]
                    if plot_cols:
                        fig4 = go.Figure()
                        colors = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728']
                        for i, col in enumerate(plot_cols):
                            fig4.add_trace(go.Scatter(x=filtered_stats['_date'], y=filtered_stats[col], mode='lines+markers', name=col, line=dict(color=colors[i % 4])))
                        fig4.update_layout(xaxis_title='日期', yaxis_title='均值', hovermode='x unified', height=400)
                        st.plotly_chart(fig4, use_container_width=True)
                else:
                    st.warning("无法计算每日统计")
            
            with tab5:
                st.subheader("💾 下载结果")
                st.write(f"原始行数: {len(df_original)} | 填充后: {len(df_filled)} | 新增: {len(df_filled)-len(df_original)}")
                excel_data = to_excel_with_highlight(df_filled, highlight_color=highlight_color)
                st.download_button("📥 下载Excel（标色）", data=excel_data, file_name="时间数据_已填充.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                csv_data = df_filled.to_csv(index=False).encode('utf-8-sig')
                st.download_button("📥 下载CSV", data=csv_data, file_name="时间数据_已填充.csv", mime="text/csv", use_container_width=True)
        else:
            st.error("分析失败，请检查配置")
    else:
        st.warning("请选择至少一个特征列")